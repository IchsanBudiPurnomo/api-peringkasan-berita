"""
auto_tuning_cached.py
=====================
Script tuning hyperparameter LexNews dengan pre-scraping cache.

Cara kerja:
  1. Scraping semua artikel SATU KALI → simpan ke artikel_cache.json
  2. Tuning parameter inti (threshold, lambda, damping) dengan semua boost = 0
  3. Tuning boost satu per satu secara KUMULATIF (hasil sebelumnya dipertahankan)
  4. Simpan hasil ke hasil_tuning.xlsx

Cara menjalankan:
  - Pastikan script ini ada di folder yang sama dengan app.py, lexrank.py, dll.
  - Jalankan Flask server dulu: python app.py
  - Lalu: python auto_tuning_cached.py
  - Jika cache sudah ada, Flask tidak perlu berjalan lagi.
"""

import os
import sys
import json
import time
import requests
import pandas as pd
from tqdm import tqdm

API_URL         = "http://127.0.0.1:5000"
DATASET_FILE    = "dataset.xlsx"
OUTPUT_FILE     = "hasil_tuning.xlsx"
CACHE_FILE      = "artikel_cache.json"
REQUEST_TIMEOUT = 60

# ---------------------------------------------------------------------------
# RANGES — rentang nilai yang akan diuji per parameter
# Nilai di luar batas masuk akal sudah dihapus.
# ---------------------------------------------------------------------------
RANGES = {
    # Core LexRank — threshold terlalu tinggi membuat graph terlalu sparse
    "SIMILARITY_THRESHOLD": [0.05, 0.08, 0.10, 0.12, 0.15, 0.18, 0.20,
                              0.25, 0.30, 0.35, 0.40, 0.45, 0.50],

    # MMR lambda — paper asli (Carbonell & Goldstein, 1998) menyatakan λ > 0.5
    # Nilai di bawah 0.5 bertentangan dengan asumsi paper, tidak diuji.
    "MMR_LAMBDA":            [0.50, 0.55, 0.60, 0.65, 0.70, 0.75, 0.80, 0.85, 0.90],

    # Damping PageRank — di bawah 0.5 tidak konvergen dengan baik (literatur: 0.75–0.90)
    "PAGERANK_DAMPING":      [0.50, 0.55, 0.60, 0.65, 0.70, 0.75,
                              0.80, 0.85, 0.90, 0.95],

    # Boost weights — 0.0 = mati, sampai 1.50 sudah lebih dari cukup
    "W_TITLE":           [0.00, 0.10, 0.20, 0.30, 0.40, 0.50, 0.60,
                          0.70, 0.80, 0.85, 0.90, 1.00, 1.20, 1.50],
    "W_POSITION":        [0.00, 0.10, 0.20, 0.25, 0.30, 0.35, 0.40,
                          0.50, 0.60, 0.70, 0.80, 1.00],
    "W_INFO_DENSITY":    [0.00, 0.05, 0.10, 0.15, 0.20, 0.25, 0.30,
                          0.40, 0.50, 0.60, 0.70, 0.80],
    "W_TITLE_INTENT":    [0.00, 0.10, 0.20, 0.25, 0.30, 0.35, 0.40,
                          0.50, 0.60, 0.70, 0.80, 1.00],
    "W_ENUMERATION":     [0.00, 0.10, 0.15, 0.20, 0.25, 0.30, 0.35,
                          0.40, 0.50, 0.60, 0.70, 0.80],
    "W_EXPERT":          [0.00, 0.10, 0.20, 0.30, 0.35, 0.40, 0.45,
                          0.50, 0.60, 0.70, 0.80, 1.00],
    "W_RECOMMENDATION":  [0.00, 0.10, 0.15, 0.20, 0.25, 0.30, 0.35,
                          0.40, 0.50, 0.60, 0.70, 0.80],
    "W_POLICY":          [0.00, 0.10, 0.20, 0.30, 0.35, 0.40, 0.45,
                          0.50, 0.60, 0.70, 0.80, 1.00],
}

# Default awal untuk parameter inti (sesuai source code asli)
DEFAULTS_INTI = {
    "SIMILARITY_THRESHOLD": 0.10,
    "MMR_LAMBDA":            0.50,   # sesuai paper: λ > 0.5, jadi 0.50 adalah batas bawah
    "PAGERANK_DAMPING":      0.85,
}

# ---------------------------------------------------------------------------
# Import modul lokal
# ---------------------------------------------------------------------------
try:
    from preprocessor import clean_text, segment_sentences
    from lexrank import score_sentences
    from mmr import select_sentences_mmr, compute_dynamic_summary_length
    from utils import reconstruct_summary
    from evaluasi import evaluate_rouge, calculate_5w1h_coverage
    import lexrank as _lexrank_mod
    import mmr as _mmr_mod
except ImportError as e:
    print(f"[ERROR] Gagal import modul: {e}")
    print("Pastikan script ini dijalankan dari folder yang sama dengan app.py!")
    sys.exit(1)


# ===========================================================================
# UTILITIES
# ===========================================================================

def print_header(title: str):
    print("\n" + "=" * 65)
    print(f"  {title}")
    print("=" * 65)


def load_dataset():
    if not os.path.exists(DATASET_FILE):
        print(f"[ERROR] File '{DATASET_FILE}' tidak ditemukan!")
        sys.exit(1)
    df = pd.read_excel(DATASET_FILE)
    url_col, ref_col = None, None
    for c in df.columns:
        cl = str(c).lower()
        if "url" in cl or "link" in cl:
            url_col = c
        if any(k in cl for k in ["reference", "ringkasan", "summary", "referensi"]):
            ref_col = c
    if not url_col or not ref_col:
        print("[ERROR] Kolom URL atau Referensi tidak ditemukan di dataset!")
        print(f"  Kolom yang ada: {list(df.columns)}")
        sys.exit(1)
    df = df.dropna(subset=[url_col, ref_col])
    df = df[df[url_col].astype(str).str.startswith("http")]
    print(f"[OK] Dataset: {len(df)} artikel | URL='{url_col}' | Ref='{ref_col}'")
    return df, url_col, ref_col


def apply_config(config: dict):
    """Terapkan nilai-nilai config ke modul lexrank dan mmr secara langsung."""
    lexrank_params = {
        "SIMILARITY_THRESHOLD", "PAGERANK_DAMPING", "TOP_TERMS_K",
        "W_POSITION", "W_TITLE", "W_INFO_DENSITY",
        "W_TITLE_INTENT", "W_ENUMERATION", "W_EXPERT",
        "W_RECOMMENDATION", "W_POLICY",
    }
    mmr_params = {"MMR_LAMBDA"}
    for param, val in config.items():
        if param in lexrank_params:
            setattr(_lexrank_mod, param, val)
        elif param in mmr_params:
            setattr(_mmr_mod, param, val)


# ===========================================================================
# FASE 1: CACHE (scraping satu kali)
# ===========================================================================

def build_cache(df, url_col, ref_col) -> list:
    """
    Scraping semua artikel satu kali dan simpan hasilnya ke JSON.
    Jika cache sudah ada, langsung pakai tanpa scraping ulang.
    Scraping dilakukan langsung (tanpa hit Flask) agar tidak bergantung
    pada parameter tuning yang sedang diubah-ubah.
    """
    if os.path.exists(CACHE_FILE):
        print(f"[OK] Cache ditemukan: '{CACHE_FILE}' — scraping dilewati.")
        with open(CACHE_FILE, "r", encoding="utf-8") as f:
            cache = json.load(f)
        print(f"[OK] {len(cache)} artikel ter-cache siap untuk tuning.")
        return cache

    print_header("FASE 1: PRE-SCRAPING ARTIKEL (hanya 1x)")
    print(f"Scraping {len(df)} URL langsung (tanpa Flask, tanpa scraping ulang).")

    from scraper import fetch_article

    cache  = []
    failed = []

    for _, row in tqdm(df.iterrows(), total=len(df), desc="Scraping"):
        url       = str(row[url_col]).strip()
        reference = str(row[ref_col]).strip()

        if not url.startswith("http"):
            continue

        try:
            article_data = fetch_article(url)
            raw_text     = article_data["text"]
            title        = article_data.get("title") or ""

            clean_article = clean_text(raw_text)
            sentences     = segment_sentences(clean_article)

            if len(sentences) >= 2:
                cache.append({
                    "url":       url,
                    "reference": reference,
                    "title":     title,
                    "sentences": sentences,
                })
            else:
                failed.append({"url": url, "alasan": "< 2 kalimat setelah segmentasi"})

            time.sleep(1.0)   # jeda ringan agar tidak kena rate-limit

        except Exception as e:
            failed.append({"url": url, "alasan": str(e)})
            time.sleep(2.0)

    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)

    print(f"\n[OK] Cache disimpan: '{CACHE_FILE}'")
    print(f"     Berhasil : {len(cache)} artikel")
    print(f"     Gagal    : {len(failed)} artikel")
    for fi in failed[:10]:   # tampilkan maks 10 error
        print(f"       - {fi['url'][:70]} → {fi['alasan']}")

    return cache


# ===========================================================================
# EVALUASI (pakai cache, tanpa scraping)
# ===========================================================================

def evaluate_rouge_cache(cache: list) -> dict:
    """Hitung rata-rata semua metrik ROUGE untuk seluruh artikel di cache."""
    buckets = {k: [] for k in [
        "rouge1_p", "rouge1_r", "rouge1_f",
        "rouge2_p", "rouge2_r", "rouge2_f",
        "rougeL_p", "rougeL_r", "rougeL_f",
    ]}

    for item in cache:
        sentences = item["sentences"]
        reference = item["reference"]
        title     = item.get("title", "")

        if not sentences or not reference:
            continue
        try:
            scores   = score_sentences(sentences, title=title or None)
            target   = compute_dynamic_summary_length(len(sentences))
            selected = select_sentences_mmr(sentences, scores, num_sentences=target)
            summary  = reconstruct_summary(selected)
            if not summary:
                continue
            r = evaluate_rouge(summary, reference)
            buckets["rouge1_p"].append(r["rouge1"]["precision"])
            buckets["rouge1_r"].append(r["rouge1"]["recall"])
            buckets["rouge1_f"].append(r["rouge1"]["fmeasure"])
            buckets["rouge2_p"].append(r["rouge2"]["precision"])
            buckets["rouge2_r"].append(r["rouge2"]["recall"])
            buckets["rouge2_f"].append(r["rouge2"]["fmeasure"])
            buckets["rougeL_p"].append(r["rougeL"]["precision"])
            buckets["rougeL_r"].append(r["rougeL"]["recall"])
            buckets["rougeL_f"].append(r["rougeL"]["fmeasure"])
        except Exception:
            continue

    n = len(buckets["rouge1_f"])
    if n == 0:
        return {k: 0.0 for k in buckets}
    return {k: round(sum(v) / n, 4) for k, v in buckets.items()}


def evaluate_5w1h_cache(cache: list) -> dict:
    """Hitung rata-rata coverage 5W1H untuk seluruh artikel di cache."""
    buckets = {k: [] for k in ["who", "what", "where", "when", "why", "how"]}

    for item in cache:
        sentences = item["sentences"]
        reference = item["reference"]
        title     = item.get("title", "")

        if not sentences or not reference:
            continue
        try:
            scores   = score_sentences(sentences, title=title or None)
            target   = compute_dynamic_summary_length(len(sentences))
            selected = select_sentences_mmr(sentences, scores, num_sentences=target)
            summary  = reconstruct_summary(selected)
            if not summary:
                continue
            cov = calculate_5w1h_coverage(summary, reference)
            for k in buckets:
                if cov.get(k) is not None:
                    buckets[k].append(cov[k])
        except Exception:
            continue

    result = {}
    for k, v in buckets.items():
        result[k] = round(sum(v) / len(v), 4) if v else 0.0
    vals = [v for v in result.values() if v is not None]
    result["average"] = round(sum(vals) / len(vals), 4) if vals else 0.0
    return result


def print_scores(rouge: dict, cov: dict, param: str, val):
    """Print metrik ROUGE + 5W1H ke terminal dengan format persis seperti evaluasi batch."""
    W = 35
    print(f"  {param} = {val}")
    print(f"  [1. ROUGE]")
    print(f"  ROUGE-1: Precision: {rouge['rouge1_p']:.4f} | Recall: {rouge['rouge1_r']:.4f} | F-Measure: {rouge['rouge1_f']:.4f}")
    print(f"  ROUGE-2: Precision: {rouge['rouge2_p']:.4f} | Recall: {rouge['rouge2_r']:.4f} | F-Measure: {rouge['rouge2_f']:.4f}")
    print(f"  ROUGE-L: Precision: {rouge['rougeL_p']:.4f} | Recall: {rouge['rougeL_r']:.4f} | F-Measure: {rouge['rougeL_f']:.4f}")
    print(f"  [2. COVERAGE 5W1H (Tingkat Ketangkapan Fakta)]")
    print(f"  {'WHO COVERAGE':<{W}}: {cov['who']*100:.2f}%")
    print(f"  {'WHAT COVERAGE':<{W}}: {cov['what']*100:.2f}%")
    print(f"  {'WHERE COVERAGE':<{W}}: {cov['where']*100:.2f}%")
    print(f"  {'WHEN COVERAGE':<{W}}: {cov['when']*100:.2f}%")
    print(f"  {'WHY COVERAGE':<{W}}: {cov['why']*100:.2f}%")
    print(f"  {'HOW COVERAGE':<{W}}: {cov['how']*100:.2f}%")
    print(f"  {'AVERAGE 5W1H COVERAGE':<{W}}: {cov['average']*100:.2f}%")
    print()


# ===========================================================================
# SIMPAN EXCEL
# ===========================================================================

def save_excel(all_results: list, optimal_config: dict,
               inti_params: list, boost_params: list,
               final_rouge: dict, final_cov: dict):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows

    df_all = pd.DataFrame(all_results)

    thin   = Border(left=Side(style="thin"),  right=Side(style="thin"),
                    top=Side(style="thin"),   bottom=Side(style="thin"))
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    DARK   = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    MID    = PatternFill("solid", start_color="2F5496", end_color="2F5496")
    HDR    = PatternFill("solid", start_color="4472C4", end_color="4472C4")
    ALT    = PatternFill("solid", start_color="DCE6F1", end_color="DCE6F1")
    BEST   = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")  # hijau muda = baris optimal

    def wf(bold=False, size=10, color="FFFFFF"):
        return Font(name="Arial", bold=bold, color=color, size=size)
    def bf(bold=False, size=10):
        return Font(name="Arial", bold=bold, size=size)

    wb = Workbook()

    # ----------------------------------------------------------------
    # Sheet 1: Semua Hasil Tuning
    # ----------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Semua Hasil"

    for r_idx, row_data in enumerate(dataframe_to_rows(df_all, index=False, header=True), start=1):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.border    = thin
            cell.alignment = center
            if r_idx == 1:
                cell.font = wf(bold=True)
                cell.fill = MID
            else:
                cell.font = bf()
                if isinstance(val, float):
                    cell.number_format = "0.0000"

    # Highlight baris optimal (ROUGE-1 F tertinggi per parameter)
    param_col = list(df_all.columns).index("parameter") + 1
    r1f_col   = list(df_all.columns).index("rouge1_f") + 1
    for param in df_all["parameter"].unique():
        sub  = df_all[df_all["parameter"] == param]
        best = sub["rouge1_f"].idxmax()
        excel_row = best + 2   # +1 header, +1 karena pandas 0-indexed
        for c_idx in range(1, len(df_all.columns) + 1):
            ws1.cell(row=excel_row, column=c_idx).fill = BEST

    ws1.freeze_panes = "A2"
    ws1.row_dimensions[1].height = 22
    for i, col in enumerate(df_all.columns, start=1):
        max_len = max(df_all[col].map(lambda x: len(str(x))).max(), len(str(col)))
        ws1.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 22)

    # ----------------------------------------------------------------
    # Sheet 2: Parameter Optimal
    # ----------------------------------------------------------------
    ws2 = wb.create_sheet("Parameter Optimal")

    KETERANGAN = {
        "SIMILARITY_THRESHOLD": "Threshold cosine sim untuk edge graph LexRank",
        "MMR_LAMBDA":            "Trade-off relevance (tinggi) vs diversity (rendah)",
        "PAGERANK_DAMPING":      "Damping factor PageRank — dari literatur 0.75–0.90",
        "W_TITLE":               "Kemiripan kalimat vs judul artikel",
        "W_POSITION":            "Kalimat awal dapat boost (inverted pyramid)",
        "W_INFO_DENSITY":        "Kepadatan angka, tanggal, kutipan, entitas",
        "W_TITLE_INTENT":        "Overlap top TF-IDF terms + sinyal urgensi",
        "W_ENUMERATION":         "Deteksi daftar bernomor dan kata ordinal",
        "W_EXPERT":              "Pernyataan narasumber/ahli dan kausalitas",
        "W_RECOMMENDATION":      "Kalimat saran, rekomendasi, atau solusi",
        "W_POLICY":              "Penetapan kebijakan, jadwal definitif",
    }

    def write_section_header(ws, row, text, fill, ncols=4):
        ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
        c = ws.cell(row=row, column=1, value=text)
        c.font = wf(bold=True, size=11); c.fill = fill
        c.alignment = center; c.border = thin
        ws.row_dimensions[row].height = 22

    row = 1
    write_section_header(ws2, row, "PARAMETER OPTIMAL HASIL TUNING", DARK, ncols=4)
    row += 2

    for group_label, param_list in [("Parameter Inti", inti_params), ("Parameter Boost", boost_params)]:
        write_section_header(ws2, row, f"[{group_label}]", MID, ncols=4)
        row += 1

        for c_idx, hdr in enumerate(["Parameter", "Nilai Optimal", "Default Awal", "Keterangan"], start=1):
            c = ws2.cell(row=row, column=c_idx, value=hdr)
            c.font = wf(bold=True); c.fill = HDR
            c.alignment = center; c.border = thin
        row += 1

        all_defaults = {**DEFAULTS_INTI,
                        "W_TITLE": 0.85, "W_POSITION": 0.35, "W_INFO_DENSITY": 0.15,
                        "W_TITLE_INTENT": 0.30, "W_ENUMERATION": 0.25,
                        "W_EXPERT": 0.40, "W_RECOMMENDATION": 0.30, "W_POLICY": 0.40}

        for i, param in enumerate(param_list):
            fill = ALT if i % 2 == 0 else None
            vals = [param,
                    optimal_config.get(param, "-"),
                    all_defaults.get(param, "-"),
                    KETERANGAN.get(param, "")]
            for c_idx, val in enumerate(vals, start=1):
                c = ws2.cell(row=row, column=c_idx, value=val)
                c.font = bf(bold=(c_idx == 1))
                c.alignment = center; c.border = thin
                if fill: c.fill = fill
            row += 1
        row += 1

    for col_letter, width in [("A", 26), ("B", 16), ("C", 14), ("D", 48)]:
        ws2.column_dimensions[col_letter].width = width

    # ----------------------------------------------------------------
    # Sheet 3: Ringkasan ROUGE + 5W1H Final
    # ----------------------------------------------------------------
    ws3 = wb.create_sheet("Hasil Akhir")

    write_section_header(ws3, 1, "PERFORMA SISTEM DENGAN PARAMETER OPTIMAL", DARK, ncols=4)
    ws3.row_dimensions[1].height = 24

    row = 3
    for c_idx, hdr in enumerate(["Metrik", "Precision", "Recall", "F-Measure"], start=1):
        c = ws3.cell(row=row, column=c_idx, value=hdr)
        c.font = wf(bold=True); c.fill = HDR
        c.alignment = center; c.border = thin
    row += 1

    rouge_rows = [
        ("ROUGE-1", final_rouge["rouge1_p"], final_rouge["rouge1_r"], final_rouge["rouge1_f"]),
        ("ROUGE-2", final_rouge["rouge2_p"], final_rouge["rouge2_r"], final_rouge["rouge2_f"]),
        ("ROUGE-L", final_rouge["rougeL_p"], final_rouge["rougeL_r"], final_rouge["rougeL_f"]),
    ]
    for i, (label, p, r, f) in enumerate(rouge_rows):
        fill = ALT if i % 2 == 0 else None
        for c_idx, val in enumerate([label, p, r, f], start=1):
            c = ws3.cell(row=row, column=c_idx, value=val)
            c.font = bf(bold=(c_idx == 1))
            c.alignment = center; c.border = thin
            if fill: c.fill = fill
            if c_idx > 1: c.number_format = "0.0000"
        row += 1

    row += 1
    write_section_header(ws3, row, "COVERAGE 5W1H", MID, ncols=4)
    row += 1
    for c_idx, hdr in enumerate(["Elemen", "Coverage (%)", "", ""], start=1):
        c = ws3.cell(row=row, column=c_idx, value=hdr if c_idx <= 2 else "")
        c.font = wf(bold=True); c.fill = HDR
        c.alignment = center; c.border = thin
    row += 1

    for i, (k, label) in enumerate([
        ("who", "Who"), ("what", "What"), ("where", "Where"),
        ("when", "When"), ("why", "Why"), ("how", "How"), ("average", "Rata-rata")
    ]):
        fill = ALT if i % 2 == 0 else None
        val_pct = round(final_cov.get(k, 0) * 100, 2)
        is_avg = (k == "average")
        for c_idx in range(1, 5):
            c = ws3.cell(row=row, column=c_idx)
            c.border = thin; c.alignment = center
            if fill: c.fill = fill
            if c_idx == 1:
                c.value = label
                c.font = bf(bold=is_avg)
            elif c_idx == 2:
                c.value = val_pct
                c.number_format = "0.00\%"
                c.font = bf(bold=is_avg)
        row += 1

    for col_letter, width in [("A", 18), ("B", 18), ("C", 10), ("D", 10)]:
        ws3.column_dimensions[col_letter].width = width

    wb.save(OUTPUT_FILE)
    print(f"[OK] Excel disimpan: '{OUTPUT_FILE}'")


# ===========================================================================
# TUNING UTAMA
# ===========================================================================

def run_tuning():
    print_header("AUTO TUNING LEXNEWS — CACHED (tanpa scraping ulang)")

    df, url_col, ref_col = load_dataset()
    cache = build_cache(df, url_col, ref_col)

    if not cache:
        print("[ERROR] Cache kosong! Tidak ada artikel yang berhasil di-cache.")
        sys.exit(1)

    print(f"\n[INFO] Tuning menggunakan {len(cache)} artikel dari cache.")

    optimal_config = {}
    all_results    = []

    ZERO_BOOST = {
        "W_TITLE": 0.0, "W_POSITION": 0.0, "W_INFO_DENSITY": 0.0,
        "W_TITLE_INTENT": 0.0, "W_ENUMERATION": 0.0, "W_EXPERT": 0.0,
        "W_RECOMMENDATION": 0.0, "W_POLICY": 0.0,
    }

    # ----------------------------------------------------------------
    # TAHAP 1: Parameter Inti (semua boost = 0)
    # Tune satu per satu secara greedy: nilai optimal setiap parameter
    # dipertahankan saat tune parameter berikutnya.
    # ----------------------------------------------------------------
    print_header("TAHAP 1: TUNING PARAMETER INTI (semua boost = 0)")

    inti_params = ["SIMILARITY_THRESHOLD", "MMR_LAMBDA", "PAGERANK_DAMPING"]

    for param in inti_params:
        print(f"\n  Tuning: {param}")
        print(f"  {'─'*60}")

        # Gunakan optimal sebelumnya + default untuk yang belum di-tune
        current = {**DEFAULTS_INTI, **optimal_config, **ZERO_BOOST}

        best_val    = current[param]   # default sebagai fallback
        best_r1f    = -1.0

        for val in RANGES[param]:
            current[param] = val
            apply_config(current)

            rouge = evaluate_rouge_cache(cache)
            cov   = evaluate_5w1h_cache(cache)
            print_scores(rouge, cov, param, val)

            all_results.append({
                "tahap": "1_Inti", "parameter": param, "nilai": val,
                **rouge,
                "5w1h_avg": cov["average"],
            })

            if rouge["rouge1_f"] > best_r1f:
                best_r1f = rouge["rouge1_f"]
                best_val = val

        optimal_config[param] = best_val
        print(f"\n  >>> OPTIMAL {param} = {best_val}  (ROUGE-1 F: {best_r1f:.4f})")

    print_header("Hasil Optimal Parameter Inti")
    for p, v in optimal_config.items():
        print(f"  {p:30s} = {v}")

    # ----------------------------------------------------------------
    # TAHAP 2: Boost Weights — KUMULATIF
    # Setiap boost yang sudah di-tune nilainya dipertahankan saat
    # tune boost berikutnya. Dimulai dari semua boost = 0.
    # ----------------------------------------------------------------
    print_header("TAHAP 2: TUNING BOOST WEIGHTS (kumulatif)")

    boost_params = [
        "W_TITLE",          # paling berpengaruh, tune duluan
        "W_POSITION",
        "W_EXPERT",
        "W_POLICY",
        "W_TITLE_INTENT",
        "W_RECOMMENDATION",
        "W_ENUMERATION",
        "W_INFO_DENSITY",   # biasanya kontribusi kecil, tune terakhir
    ]

    # Mulai dari semua boost = 0, tambahkan nilai optimal secara kumulatif
    boost_optimal = dict(ZERO_BOOST)

    for param in boost_params:
        print(f"\n  Tuning Boost: {param}")
        print(f"  (boost aktif sebelumnya: { {k:v for k,v in boost_optimal.items() if v > 0} })")
        print(f"  {'─'*60}")

        best_val = 0.0   # fallback: matikan boost ini jika tidak ada yang lebih baik
        best_r1f = -1.0

        for val in RANGES[param]:
            current = {**optimal_config, **boost_optimal, param: val}
            apply_config(current)

            rouge = evaluate_rouge_cache(cache)
            cov   = evaluate_5w1h_cache(cache)
            print_scores(rouge, cov, param, val)

            all_results.append({
                "tahap": "2_Boost", "parameter": param, "nilai": val,
                **rouge,
                "5w1h_avg": cov["average"],
            })

            if rouge["rouge1_f"] > best_r1f:
                best_r1f = rouge["rouge1_f"]
                best_val = val

        # Simpan nilai optimal dan lanjutkan ke boost berikutnya
        boost_optimal[param] = best_val
        optimal_config[param] = best_val
        print(f"\n  >>> OPTIMAL {param} = {best_val}  (ROUGE-1 F: {best_r1f:.4f})")

    # ----------------------------------------------------------------
    # Evaluasi Final dengan seluruh config optimal
    # ----------------------------------------------------------------
    print_header("EVALUASI FINAL — Config Optimal Lengkap")
    apply_config(optimal_config)
    final_rouge = evaluate_rouge_cache(cache)
    final_cov   = evaluate_5w1h_cache(cache)

    print("\nParameter Inti:")
    for p in inti_params:
        print(f"  {p:30s} = {optimal_config[p]}")
    print("\nParameter Boost:")
    for p in boost_params:
        print(f"  {p:30s} = {optimal_config[p]}")

    print("\n[1. ROUGE]")
    print(f"  ROUGE-1: P={final_rouge['rouge1_p']:.4f}  R={final_rouge['rouge1_r']:.4f}  F={final_rouge['rouge1_f']:.4f}")
    print(f"  ROUGE-2: P={final_rouge['rouge2_p']:.4f}  R={final_rouge['rouge2_r']:.4f}  F={final_rouge['rouge2_f']:.4f}")
    print(f"  ROUGE-L: P={final_rouge['rougeL_p']:.4f}  R={final_rouge['rougeL_r']:.4f}  F={final_rouge['rougeL_f']:.4f}")

    print("\n[2. COVERAGE 5W1H]")
    W = 30
    for k in ["who", "what", "where", "when", "why", "how", "average"]:
        print(f"  {k.upper()+' COVERAGE':<{W}}: {final_cov[k]*100:.2f}%")

    # ----------------------------------------------------------------
    # Simpan ke Excel
    # ----------------------------------------------------------------
    print_header("MENYIMPAN KE EXCEL")
    save_excel(all_results, optimal_config, inti_params, boost_params,
               final_rouge, final_cov)

    print(f"\n[SELESAI]")
    print(f"  Excel  : {OUTPUT_FILE}")
    print(f"  Cache  : {CACHE_FILE}  (simpan — tidak perlu scraping ulang)")


if __name__ == "__main__":
    run_tuning()